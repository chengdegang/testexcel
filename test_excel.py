import shutil
from _csv import writer
from pathlib import Path
import openpyxl
import xlrd
import time
from datetime import datetime
from openpyxl.cell import MergedCell
import os
from openpyxl.workbook import Workbook
import pandas as pd

file1 = ''
file_xls = '/Users/jackrechard/PycharmProjects/testexcel/file/xxx.xls'
file_xlsx = '/Users/jackrechard/PycharmProjects/testexcel/file/xxx2.xlsx'

"""
支持xlsx格式的读,输入要读的文件路径
"""
def openpy_read_xlsx(file,sheetname):
    wb = openpyxl.load_workbook(file)
    table = wb[sheetname]
    #获取行数列数
    nrow = table.max_row
    ncol = table.max_column
    need_data2 = []

    #获取标题
    for row in range(1,nrow+1):
        for col in range(1,ncol+1):
            #如果这一行中存在title字样，遍历这一行的所有数据，添加到列表中
            if "title" in str(table.cell(row, col).value):
                for i in range(1,ncol+1):
                    # print(table.cell(row, i).value)
                    need_data2.append(table.cell(row, i).value)
                # print(need_data2)
                break

    #获取需要的行，判断条件为包含'_1'
    for row in range(1,nrow+1):
        for col in range(1,ncol+1):
            #判断是否包含'value'，是的话选取整列
            if "value" in str(table.cell(row,col).value):
                #判断这列是否包含'_1'，是的话取整行
                # print(table.cell(row, col).value)
                #遍历这一列的所有数据，并标记存在'_1'的行
                for i in range(1,nrow+1):
                    if "_1" in str(table.cell(i,col).value):
                        # print(table.cell(i,col).value
                        #将找到的这一行所有数据写入列表
                        for j in range(1,ncol+1):
                            # print(table.cell(i,j).value)
                            need_data2.append(table.cell(i,j).value)

    #将改列表按照列数切成多个列表
    per_list_len = ncol
    list_of_group = zip(*(iter(need_data2),) * per_list_len)
    end_list = [list(i) for i in list_of_group]  # i is a tuple
    count = len(need_data2) % per_list_len
    end_list.append(need_data2[-count:]) if count != 0 else end_list
    print(end_list)
    return end_list

"""
写入均只支持xlsx，两种写法，用第二种即可
"""
#写法1，列为手动写的，有长度限制
# data = [['title1', 'title2', 'title_value3'], ['数据4', '数据5', '数据6_1'], ['数据10', '数据11', '数据12_1']]
def write_excel_(file,data):
    #内存中创建一个空表格
    # wb = openpyxl.Workbook()
    # sheet = wb.active
    # sheet.title = 'ccc_sheet'
    # print(wb.sheetnames)

    #写到原来的表格中，新建一个sheet,只支持xlsx文件
    wb = openpyxl.load_workbook(file)
    wb.create_sheet(index=0,title='ccc_sheet')
    sheet = wb['ccc_sheet']
    abc = "0ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    for i in range(len(data)):
        for j in range(len(data[i])):
            sheet[abc[j+1]+str(i+1)] = data[i][j]
            # sheet.cell(row=i+1,column=j+1,value=data[i][j])
            print(f'字母是{abc[j+1]} 数字是{i+1} 数据是{data[i][j]}')
    wb.save(file)

data2 = [['John Brown', 18, 'New York No. 1 Lake Park'],['John Brown2', 11, 'New York No. 1 Lake Park2']]
#写法2，较好的写法
"""
在当前路径创建xlsx文件,若已存在会直接覆盖写入数据
"""
def write_excel(file = '默认.xlsx',data = [['默认数据3'],['默认数据5']],sheetname = 'ces'):
    if os.path.exists(file) == True:
        wb = openpyxl.load_workbook(file)
        # 默认写到第一个sheet中，index为0
        wb.create_sheet(index=0, title=f'{sheetname}')
        sheet = wb[f'{sheetname}']
        # 好的写法写入excel
        for row_index, row_item in enumerate(data):
            for col_index, col_item in enumerate(row_item):
                sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item)
        wb.save(file)
        print('---写入完成---')
    else:
        wbc = Workbook()
        log_path = os.getcwd() + '/'
        # t = time.strftime('%Y%m%d_%H%M', time.localtime(time.time()))
        # suffix = '.xlsx'  # 文件类型
        # newfile = file + suffix
        path = log_path + file
        wbc.save(path)
        print(f"创建文件 {log_path + file} ")
        wb = openpyxl.load_workbook(log_path + file)
        # 默认写到第一个sheet中，index为0
        wb.create_sheet(index=0, title=f'{sheetname}')
        sheet = wb[f'{sheetname}']
        # 好的写法写入excel
        for row_index, row_item in enumerate(data):
            for col_index, col_item in enumerate(row_item):
                sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item)
        wb.save(log_path + file)
        print('---写入完成---')
        return log_path + file

def merged_deal_xlsx(file):
    """
    处理xlsx的excel合并单元格,获取文件路径，重读excel文件并输出每一个单元格重写了数据的列表（主要针对有合并单元格的情况），并去重
    """
    newdata = []
    wb = openpyxl.load_workbook(file)
    table = wb["Sheet2"]
    nrow = table.max_row
    ncol = table.max_column

    for row_index in range(1, nrow + 1):
        for col_index in range(1, ncol + 1):
            cell = table.cell(row=row_index, column=col_index)
            if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
                for merged_range in table.merged_cell_ranges:  # 循环查找该单元格所属的合并区域
                    if cell.coordinate in merged_range:
                        # 获取合并区域左上角的单元格作为该单元格的值返回
                        cell_ = table.cell(row=merged_range.min_row, column=merged_range.min_col)
                        newdata.append(cell_.value)
                        break
            else:
                cell_ = table.cell(row=row_index, column=col_index)
                # newdata.append(cell_.value)
                newdata.append(cell_.value)

    per_list_len = ncol
    list_of_group = zip(*(iter(newdata),) * per_list_len)
    end_list = [list(i) for i in list_of_group]  # i is a tuple
    count = len(newdata) % per_list_len
    end_list.append(newdata[-count:]) if count != 0 else end_list

    #去重
    end_list2 = []
    for element in end_list:
        if element not in end_list2:
            end_list2.append(element)
    print(end_list2)
    return end_list2

"""
查找指定文件夹下的指定文件，移动到指定路径并重命名
"""
def osmake():
    t = time.strftime("%Y_%m_%d_%H_%M", time.localtime())
    for f_name in os.listdir('file'):
        # print(f_name)
        # if f_name.endswith('.txt'):
        #     print(f_name)
        if '123' in f_name:
            newdir = Path(f'{os.getcwd()}/file/backups/')
            if newdir.exists():
                print('cunz')
            else:
                newdir = os.mkdir(f'{os.getcwd()}/file/backups/')

            print(f'find path: {os.getcwd()}/file/{f_name}')
            print(f'goal path: {newdir}/')
            #move
            shutil.move(str(f'{os.getcwd()}/file/{f_name}'),f'{newdir}/')
            #rename
            shutil.move(f'{newdir}/{f_name}',f"{newdir}/{f_name.split('.')[0]}_{t}.{f_name.split('.')[1]}")

"""
读文件并统计sheet及sheet中的数据量（行）
"""
def read_xlsx_1(file):
    wb = openpyxl.load_workbook(file)
    #获取所有sheet
    sheets = wb.sheetnames
    print(sheets)
    for i in range(len(sheets)):
        table = wb[sheets[i]]
        nrow = table.max_row
        print(f'{sheets[i]}行数为 : {nrow}')

def panda(file):
    f = pd.read_excel(file,sheet_name=None)
    for sheet in list(f):
        print(sheet)
        # 读取整个sheet的数据为一个矩阵
        # print(f.head())
        excel = pd.read_excel(file, sheet_name=sheet)
        print(excel.head())
    #行数
    # print(len(f.index.values))
    # print(f.index.values)
    #列数
    # print(len(f.columns.values))
    # print(f.columns.values)

    #读取第一行数据
    # data = f.ix[0].values
    # print(data)

def panda_write(file):
    df1 = pd.DataFrame(
        {'日期': [datetime(2020, 1, 1), datetime(2020, 1, 2)],
         '销量': [10, 20]}
    )
    df2 = pd.DataFrame(
        {'日期': [datetime(2020, 2, 1), datetime(2020, 2, 2)],
         '销量': [15, 25]}
    )
    with pd.ExcelWriter(
            file,
            datetime_format='YYYY-MM-DD'  # 只显示年月日, 不显示时分秒
    ) as writer:
        df1.to_excel(writer, sheet_name='1月')  # Sheet1
        df2.to_excel(writer, sheet_name='2月')  # Sheet2

"""
传入路径，遍历路径下所有xls并将其转换为xlsx
"""
def xls_to_xlsx(fdir):

    for root,dirs,files in os.walk(fdir):
        for name in files:#遍历文件夹下的文件名
            if 'xls' in name:
                # print(os.path.join(root, name))
                fname = name.split('.')
                # print(fname[0])
                excel = pd.read_excel(os.path.join(root, name))
                data = pd.DataFrame(excel)
                data.to_excel(os.path.join(root, fname[0]) + '.xlsx',index=False)

    print('转换完成')

        # for name in dirs:#遍历文件夹下的路径
        #     print(os.path.join(root,name))
    # for i in os.listdir(fdir):
    #     print(i)

# xls_to_xlsx('/Users/jackrechard/PycharmProjects/crawl_syb/data/')

# panda('/Users/jackrechard/PycharmProjects/testexcel/file/change/ces.xls')
# panda_write('/Users/jackrechard/PycharmProjects/testexcel/file/change/ces.xls')
# read_xlsx_1(file_xlsx)
# osmake()

#读xls的excel
# xlrd_read_xls(file_xxx)

#读xlsx的excel
# openpy_read_xlsx(file=file_xxx2,sheetname='ccc_sheet_ces')
#将xls筛选的数据写入到excel中
# write_excel_(file_xxx2,xlrd_read_xls(file_xxx))
#将xlsx取出的数据写入excel
# write_excel(file_xxx2,openpy_read_xlsx(file_xxx2))

# datac = [['A1324','iPhone 3G (国行)'], ['A1303','iPhone 3GS'], ['A1325','iPhone 3GS (国行, 无 WLAN 功能)']]
# data=['A1324,iPhone 3G (国行)', 'A1303,iPhone 3GS']
# datat = [['A1324', 'iPhone 3G (国行)'], ['A1303', 'iPhone 3GS']]
# write_excel(data=datat)

# write_excel(file='file/20210722193722.xlsx',data = [['默认数据1'],['默认数据2']],sheetname = 'ces2')

#读合并单元格
# merged_deal_xlsx(file_xxx2)

#     write_excel(file= file_xxx2,data= merged_deal_xlsx(file_xxx2))

