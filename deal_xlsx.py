import openpyxl
import xlrd
import time
from openpyxl.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from test_excel import write_excel

file1 = ''
file_xxx2 = '/Users/jackrechard/PycharmProjects/testexcel/file/xxx2.xlsx'

"""
支持xlsx格式的读,输入要读的文件路径,
"""
def openpy_read_xlsx(file,sheetname):
    wb = openpyxl.load_workbook(file)
    table = wb[sheetname]
    #获取行数列数
    nrow = table.max_row
    ncol = table.max_column
    need_data2 = []

    #获取标题
    for row in range(1,nrow+1):
        for col in range(1,ncol+1):
            #如果这一行中存在title字样，遍历这一行的所有数据，添加到列表中
            if "title" in str(table.cell(row, col).value):
                for i in range(1,ncol+1):
                    # print(table.cell(row, i).value)
                    need_data2.append(table.cell(row, i).value)
                # print(need_data2)
                break

    #获取需要的行，判断条件为包含'_1'
    for row in range(1,nrow+1):
        for col in range(1,ncol+1):
            #判断是否包含'value'，是的话选取整列
            if "value" in str(table.cell(row,col).value):
                #判断这列是否包含'_1'，是的话取整行
                # print(table.cell(row, col).value)
                #遍历这一列的所有数据，并标记存在'_1'的行
                for i in range(1,nrow+1):
                    if "_1" in str(table.cell(i,col).value):
                        # print(table.cell(i,col).value
                        #将找到的这一行所有数据写入列表
                        for j in range(1,ncol+1):
                            # print(table.cell(i,j).value)
                            need_data2.append(table.cell(i,j).value)

    #将改列表按照列数切成多个列表
    per_list_len = ncol
    list_of_group = zip(*(iter(need_data2),) * per_list_len)
    end_list = [list(i) for i in list_of_group]  # i is a tuple
    count = len(need_data2) % per_list_len
    end_list.append(need_data2[-count:]) if count != 0 else end_list
    print(end_list)
    return end_list

def merged_deal_xlsx(file):
    """
    处理xlsx的excel合并单元格,获取文件路径，重读excel文件并输出每一个单元格重写了数据的列表（主要针对有合并单元格的情况），并去重
    """
    newdata = []
    wb = openpyxl.load_workbook(file)
    table = wb["Sheet2"]
    nrow = table.max_row
    ncol = table.max_column

    for row_index in range(1, nrow + 1):
        for col_index in range(1, ncol + 1):
            cell = table.cell(row=row_index, column=col_index)
            if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
                for merged_range in table.merged_cell_ranges:  # 循环查找该单元格所属的合并区域
                    if cell.coordinate in merged_range:
                        # 获取合并区域左上角的单元格作为该单元格的值返回
                        cell_ = table.cell(row=merged_range.min_row, column=merged_range.min_col)
                        newdata.append(cell_.value)
                        break
            else:
                cell_ = table.cell(row=row_index, column=col_index)
                # newdata.append(cell_.value)
                newdata.append(cell_.value)

    per_list_len = ncol
    list_of_group = zip(*(iter(newdata),) * per_list_len)
    end_list = [list(i) for i in list_of_group]  # i is a tuple
    count = len(newdata) % per_list_len
    end_list.append(newdata[-count:]) if count != 0 else end_list

    #去重
    end_list2 = []
    for element in end_list:
        if element not in end_list2:
            end_list2.append(element)
    print(end_list2)
    return end_list2

#读xls的excel
# xlrd_read_xls(file_xxx)

#读xlsx的excel
# openpy_read_xlsx(file=file_xxx2,sheetname='ccc_sheet_ces')
#将xls筛选的数据写入到excel中
# write_excel_(file_xxx2,xlrd_read_xls(file_xxx))
#将xlsx取出的数据写入excel
# write_excel(file_xxx2,openpy_read_xlsx(file_xxx2))

#读合并单元格
# merged_deal_xlsx(file_xxx2)
# write_excel(file= file_xxx2,data= merged_deal_xlsx(file_xxx2))


#第一步，读取xlsx的file文件，并将其合并补全每一个合并单元格的数据，返回list，并写入到xlsx文件
#第二步，读取第一步保存的文件并筛选指定的数据返回一个list
#第三步，将这个list数据写入到文件中，命名一个sheetname
write_excel(file=file_xxx2,data=merged_deal_xlsx(file_xxx2),sheetname='cccxlsx')
data = openpy_read_xlsx(file=file_xxx2,sheetname='ccc_sheet_cccxlsx')
write_excel(file=file_xxx2,data=data,sheetname='ccc2xlsx')

